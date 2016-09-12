from openpyxl import Workbook
from dwarfsquad.lib.utils import to_stderr
from dwarfsquad.lib.export.export_lots_and_levels import build_lot_rows
from dwarfsquad.lib.export.export_rulesettings import get_rulesettings_rows, get_rulesettings_keys
from dwarfsquad.lib.export.export_compounds import build_compound_rows
from dwarfsquad.model.AssayConfiguration import AssayConfiguration
from dwarfsquad.lib.export.export_assay_configuration import get_assay_configuration_rows


def export_full_ac(ac):
    assert isinstance(ac, AssayConfiguration)
    wb = get_workbook(ac)
    wb.save(ac.name + ".xlsx")
    to_stderr("Wrote " + ac.name + ".xlsx")


def set_all_cells_to_text(ws):
    for row in ws.rows:
        for cell in row:
            cell.number_format = '@'


def get_workbook(ac):
    wb = Workbook()
    assay_worksheet = wb.create_sheet(title="Assay")
    assay_rows = get_assay_configuration_rows(ac)
    assay_worksheet.append([k for k in assay_rows[0].keys()])
    for row in assay_rows:
        assay_worksheet.append([r for r in row.values()])
    compound_worksheet = wb.create_sheet(title="Compound")
    compound_rows = build_compound_rows(ac)
    compound_worksheet.append([k for k in compound_rows[0].keys()])
    [compound_worksheet.append([r for r in row.values()]) for row in compound_rows]
    lots_levels_worksheet = wb.create_sheet(title="Lots")
    lot_rows = build_lot_rows(ac)
    lots_levels_worksheet.append([k for k in lot_rows[0].keys()])
    [lots_levels_worksheet.append([r for r in row.values()]) for row in lot_rows]
    set_all_cells_to_text(lots_levels_worksheet)
    rulesettings_worksheet = wb.create_sheet(title="Rule")
    rulesettings_rows = get_rulesettings_rows(ac)
    rulesettings_keys = get_rulesettings_keys(rulesettings_rows)
    rulesettings_worksheet.append(rulesettings_keys)
    [rulesettings_worksheet.append([r.get(k) for k in rulesettings_keys]) for r in rulesettings_rows]
    wb.remove_sheet(wb.worksheets[0])
    return wb
